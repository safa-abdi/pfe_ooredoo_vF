import datetime
import openpyxl
import re
import requests
from datetime import datetime,timedelta
import mysql.connector
from mysql.connector import Error
from collections import defaultdict

file_path = './data to clean/plaintes.xlsx'
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

def remove_duplicates(data_list):
    """Supprime les doublons basés sur CRM_CASE en gardant le DERNIER occurrence"""
    seen = {}
    for idx, data in enumerate(data_list):
        seen[data['CRM_CASE']] = idx
    
    unique_data = [data_list[idx] for idx in seen.values()]
    return unique_data

def connect_to_db():
    try:
        connection = mysql.connector.connect(
            host='localhost',
            database='ooredoo',
            user='root',
            password='#Safa@123_#'
        )
        return connection
    except Error as e:
        print(f"Erreur de connexion à la base de données : {e}")
        return None
    
def check_existing_crms(connection, crm_cases):
    cursor = connection.cursor()
    placeholders = ', '.join(['%s'] * len(crm_cases))
    query = f"SELECT CRM_CASE FROM plainte WHERE CRM_CASE IN ({placeholders})"
    cursor.execute(query, crm_cases)
    results = cursor.fetchall()
    return set(row[0] for row in results)
def insert_new_data(connection, data_to_insert):
    cursor = connection.cursor()

    insert_query = """
    INSERT INTO plainte (
        CRM_CASE, DATE_CREATION_CRM, LATITUDE_SITE, LONGITUDE_SITE, 
        MSISDN, CONTACT_CLIENT, CONTACT2_CLIENT, CLIENT, REP_TRAVAUX_STT,
        NAME_STT, Delegation, Gouvernorat, DATE_AFFECTATION_STT, DES_PACK,
        offre, OPENING_DATE_SUR_TIMOS, REP_RDV, DATE_PRISE_RDV, CMT_RDV,
        Detail, STATUT, DATE_FIN_TRV, Description
    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """

    existing_crms = set()
    if data_to_insert:
        crm_cases = [entry['CRM_CASE'] for entry in data_to_insert if entry['CRM_CASE']]
        existing_crms = check_existing_crms(connection, crm_cases)

    new_records = []
    for entry in data_to_insert:
        crm_case = entry['CRM_CASE']
        if crm_case and crm_case not in existing_crms:
            new_records.append((
                crm_case,
                entry['DATE_CREATION_CRM'],
                entry['LATITUDE_SITE'],
                entry['LONGITUDE_SITE'],
                entry['MSISDN'],
                entry['CONTACT_CLIENT'],
                entry['CONTACT2_CLIENT'],
                entry['CLIENT'],
                entry['REP_TRAVAUX_STT'],
                entry['NAME_STT'],
                entry['Delegation'],
                entry['Gouvernorat'],
                entry['DATE_AFFECTATION_STT'],
                entry['DES_PACK'],
                entry['offre'],
                entry['OPENING_DATE_SUR_TIMOS'],
                entry['REP_RDV'],
                entry['DATE_PRISE_RDV'],
                entry['CMT_RDV'],
                entry['Detail'],
                entry['STATUT'],
                entry['DATE_FIN_TRV'],
                entry['Description']
            ))

    if new_records:
        cursor.executemany(insert_query, new_records)
        connection.commit()
        print(f"{len(new_records)} nouveaux enregistrements insérés.")
    else:
        print("Aucun nouvel enregistrement à insérer.")

def generate_sql_insert(data):
    columns = [
        'CRM_CASE', 'DATE_CREATION_CRM', 'LATITUDE_SITE', 'LONGITUDE_SITE', 
        'MSISDN', 'CONTACT_CLIENT', 'CONTACT2_CLIENT', 'CLIENT', 'REP_TRAVAUX_STT',
        'NAME_STT', 'Delegation', 'Gouvernorat', 'DATE_AFFECTATION_STT', 'DES_PACK',
        'offre', 'OPENING_DATE_SUR_TIMOS', 'REP_RDV', 'DATE_PRISE_RDV', 'CMT_RDV',
        'Detail', 'STATUT', 'DATE_FIN_TRV', 'Description'
    ]
    
    values = []
    for col in columns:
        val = data.get(col)

        if val is None or val == 'NULL':
            values.append('NULL')
        elif col in ['DATE_CREATION_CRM', 'DATE_AFFECTATION_STT', 'OPENING_DATE_SUR_TIMOS', 
                     'DATE_PRISE_RDV', 'DATE_FIN_TRV'] and val:
            values.append(f"'{val}'")
        elif isinstance(val, str):
            val_escaped = val.replace("'", "''")
            values.append(f"'{val_escaped}'")
        else:
            values.append(str(val))

    sql = f"""
    INSERT INTO plainte (
        {', '.join(columns)}
    ) VALUES (
        {', '.join(values)}
    );
    """
    return sql

# Génération du fichier SQL
sql_file_path = './insert_plaintes.sql'


def verif_rep_STT(rep):
    return rep if rep else "en_cours" 

def verifierDelegation(deleg):
    #print("delegation",deleg)
    if deleg is None:
        return None
    return None if "@" in deleg else deleg
def verifierGouv(gouv):
    if(gouv):
        return gouv if gouv.isalpha() else None
    return None
def verif_Nom_STT(stt):
    timestamp_pattern = r"\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}" 
    matchTime= re.search(timestamp_pattern, str(crm_case))
    STT = str(stt)
    parts = STT.split(' ') 
    if (len(parts)==2):
        if(matchTime):
                return None 
    return stt

def verif_date(dateRDV):
    if not dateRDV or dateRDV in ["NULL", "None", ""]:
        return None
    
    # Convertir en datetime si c'est une chaîne
    if isinstance(dateRDV, str):
        dateRDV = datetime.strptime(dateRDV, "%Y-%m-%d %H:%M:%S")
    
    # Vérifier si l'heure est 00:00:00
    if dateRDV.hour == 0 and dateRDV.minute == 0 and dateRDV.second == 0:
        dateRDV = dateRDV.replace(hour=23, minute=59, second=0)
    
    return dateRDV.strftime("%Y-%m-%d %H:%M:%S")

def verif_affectation(nomSTT, dateAff, REP_TRAVAUX_STT):
    #print("stt",nomSTT)
    if not nomSTT and not dateAff and not REP_TRAVAUX_STT:
        return "non_affecté", None, "non_affecté_stt"
    return nomSTT, dateAff, REP_TRAVAUX_STT

def verif_dates(OPENING_DATE_SUR_TIMOS, DATE_PRISE_RDV, DATE_FIN_TRV):
    def parse_date(date_val):
        if date_val in [None, "NULL", "", "null"]:
            return None
        if isinstance(date_val, str):
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                try:
                    return datetime.strptime(date_val, fmt)
                except ValueError:
                    continue
        return date_val if isinstance(date_val, datetime) else None

    opening_date = parse_date(OPENING_DATE_SUR_TIMOS)
    date_prise_rdv = parse_date(DATE_PRISE_RDV)
    date_fin_trv = parse_date(DATE_FIN_TRV)
    sysdate = datetime.now()

    def adjust_future_date(date_to_adjust, reference_date):
        if reference_date.date() == sysdate.date():  # Même jour
            return sysdate + timedelta(hours=2)
        else:
            return reference_date + timedelta(days=2)

    def adjust_date(date_to_check, reference_date):
        if not date_to_check or not reference_date:
            return None

        # Vérification futur
        if date_to_check > sysdate:
            return adjust_future_date(date_to_check, reference_date)

        # Logique normale d'ajustement
        if date_to_check.year < reference_date.year:
            date_to_check = date_to_check.replace(year=reference_date.year)

        if date_to_check.year == reference_date.year:
            if date_to_check.month < reference_date.month:
                date_to_check = date_to_check.replace(month=reference_date.month)
            elif date_to_check.month == reference_date.month:
                if date_to_check.day < reference_date.day:
                    date_to_check += timedelta(days=2)

        return date_to_check if date_to_check <= sysdate else adjust_future_date(date_to_check, reference_date)

    if opening_date:
        date_prise_rdv = adjust_date(date_prise_rdv, opening_date)
        date_fin_trv = adjust_date(date_fin_trv, opening_date)

    return date_prise_rdv, date_fin_trv


def separate_numbers(input_str):
    match = re.match(r"(\d+\.\d+)\.(\d+\.\d+)", input_str)
    match100 = re.match(r"(\d+\.\d+)\_(\d+\.\d+)", input_str)
    match200 = re.match(r"(\d+\.\d+)", input_str)
    
    if match200:
        return match200.group(1),None 
    if match :
        #print("match",match)
        return match.group(1), match.group(2)  

    if match100 :
        #print("match",match)
        return match100.group(1), match100.group(2)  

    else:
        return input_str, None  

def clean_coordinate(coord):
    coord = str(coord)
    match = re.match(r"(\d+)°(\d+\.\d+)\s*([NSEWZ])", coord)
    match2 = re.match(r'^\d+\.\d{5,}\.\d{2}$', coord)

    # Regex pour capturer séparément les parties souhaitées
    pattern = r"(\d{1,3})\.(\d{1,3})°(\d{1,3}\.\d+)\s*[NSEWZ]"
    pattern0 = r"^(\d+)\.(\d+)°[NSEWZ]$"
    pattern00 = r"^(\d+)\.(\d+)°$"
    pattern2 = r"(\d{1,3}\.\d+)[°]?(\d{1,3}\.\d+)"
    
    pattern3 = r"^(\d+)\.(\d+)\.(\d+)$"
    pattern4 = r"^(\d+)\.\.(\d+)$"
    pattern5 = r"(\d+),(\d+)°([NSEZ])"
    pattern6 = r"^(\d+)°(\d+) ([NSEZ])$"
    pattern7 = r"^(\d+)°(\d+)\.(\d+) ([NSEZ])$"
    pattern8 = r"^LAT(\d+)°(\d+) (\d+\.\d+) ([NSEZ])$"
    pattern9 = r"^LO°N(\d+) (\d+) (\d+\.\d+) ([NSEZ])$"
    pattern10 = r"^(\d+\.\d+)\.$"
    pattern11 = r"^(\d+)\.(\d+)°(\d+) ([NSEZ])$"
    pattern12 = r"^(\d+)\.(\d+)\.(\d+)°([NSEZ])$"
    pattern13 = r"(\d+)°(\d+)'(\d+\.\d+)$"
    pattern14 = r"(\d+)°(\d+)'(\d+\.\d+)([NSWE])"
    pattern15 = r"^(\d+)°(\d+)\.(\d+)'([NSWE])$"
    pattern16 = r"(\d+)\)(\d+)°(\d+\.\d+) ([NSWE])"
    pattern17 = r"(\d+)°(\d+) \.(\d+\.\d+) ([NSWE])"
    pattern18 = r"(\d+)°\.(\d+) (\d+\.\d+) ([NSWE])"
    pattern19 = r"(\d+)°(\d+) (\d+\.\d+) ([NSWE])"
    pattern20 = r"(\d+)° (\d+)'(\d+\.\d+)\"([NSWE])"
    pattern21 = r"^(\d+)°(\d+) (\d+) (\d+)\.(\d+) ([NSWE])$"
    pattern22 = r"(\d+)°(\d+) (\d+\.\d+)\. ([NSWE])"
    pattern23 = r"(\d+)° (\d+)'(\d+\.\d+)$"
    pattern24 = r"(\d+\.\d+),(\d+\.\d+)"
    pattern25 = r"(\d+)\,(\d+)\/(\d+)\,(\d+)"
    pattern26 = r"(\d{1,2})°(\d{1,2}) (\d{1,2}\.\d+)\* ([NSEW])"
    pattern27 = r"(\d+\.\d+)\+(\d+)"
    pattern28 = r"(\d+)°(\d+)'(\d+\.\d+)\" ([NSWE])"
    pattern29 = r"(\d+\.\d+)°\s*(\d+)'?"
    pattern30 = r"^(\d+)°(\d+)\.(\d+)\.(\d+) ([NSEZ])$"
    pattern31 = r"^(\d+)°(\d+) \.(\d+)\. ([NSEZ])$"
    pattern32 = r"^(\d+)°(\d+) (\d+) \.(\d+) ([NSEZ])$"
    pattern33 = r"^(\d+)\.(\d+)\.(\d+)\.°([NSEZ])$"
    pattern34 = r"^(\d+)\.(\d+)\.(\d+)([NSEZF])$"
    pattern35 = r"^(\d+)\,(\d+)°$"
    pattern36 =r"^([A-Za-z]+)(\d+\.\d+)"
    pattern37 = r"^(\d+)\.(\d+)\.(\d+)°$"
    pattern38 = r"^(\d+)\.(\d+)\.(\d+)\.(\d+)$"
    pattern39 = r"^(3[1-7]|1[01])(\d+)$"
    pattern40 = r"^(\d)\.(\d+)°$"
    pattern43 = r"^(\d+)°(\d+) (\d+\.\d+) $"
    pattern44 = r"^(\d+)\.(\d+)$"
    pattern45 = r"^(\d+)\.(\d+)\.(\d+)\.(\d+)\.(\d+)$"
    pattern46 = r"^(\d+)°(\d+)\?(\d+)\.(\d+)\?\? ([NSEZF])$"
    pattern47 = r"[-+]?\d{1,2}\.\d+"
    pattern48 = r"(\d+)°(\d+)'(\d+\.\d+)\"([NSEZF])$"
    match00 =re.search(pattern00,coord)
    match45 = re.search(pattern45, coord)
    match44 = re.search(pattern44, coord)
    match43 = re.search(pattern43, coord)
    # Vérifier le match
    match3 = re.search(pattern, coord)
    match4 = re.search(pattern2, coord)
    match5 = re.search(pattern3, coord)
    match6 = re.search(pattern4, coord)
    match7 = re.search(pattern5, coord)
    match8 = re.search(pattern6, coord)
    match9 = re.search(pattern7, coord)
    match10 = re.search(pattern8, coord)
    match11 = re.search(pattern9, coord)
    match12 = re.search(pattern10, coord)
    match13 = re.search(pattern11, coord)
    match14 = re.search(pattern12, coord)
    match15 = re.search(pattern13, coord)
    match16 = re.search(pattern14, coord)
    match17 = re.search(pattern15, coord)
    match18 = re.search(pattern16, coord)
    match19 = re.search(pattern17, coord)
    match20 = re.search(pattern18, coord)
    match21 = re.search(pattern19, coord)
    match22 = re.search(pattern20, coord)
    match23 = re.search(pattern21, coord)
    match24 = re.search(pattern22, coord)
    match25 = re.search(pattern23, coord)
    match26 = re.search(pattern24, coord)
    match27 = re.search(pattern25, coord)
    match28 = re.search(pattern26, coord)
    match29 = re.search(pattern27, coord)
    match30 = re.search(pattern28, coord)
    match31 = re.search(pattern29, coord)
    match32 = re.search(pattern30, coord)
    match33 = re.search(pattern31, coord)
    match34 = re.search(pattern32, coord)
    match35 = re.search(pattern33, coord)
    match36 = re.search(pattern34, coord)
    match37 = re.search(pattern35, coord)
    match38 = re.search(pattern36, coord)
    match39 = re.search(pattern37, coord)
    match40 = re.search(pattern38, coord)
    match41 = re.match(pattern39, str(coord))
    match0 = re.match(pattern0, str(coord))
    match42 = re.match(pattern40, str(coord))
    match46 = re.match(pattern46, str(coord))
    match47 = re.match(pattern47, str(coord))
    match48 = re.match(pattern48, str(coord))

    if(match00):
        nb1 = match00.group(1)
        nb2 = match00.group(2)
        nb20= int(nb2[0])
        nb21= int(nb2[1])
        if(int(nb1)==0):
            if(nb20 == 1 and (nb21 == 0 or nb21 ==1)):
                coord=nb2[0]+nb2[1]+'.'+nb2[2:]
            elif(nb20 == 3 and (nb21 > 0 and nb21 < 8)):
                coord=nb2[0]+nb2[1]+'.'+nb2[2:]
        #print("coord",coord)

    if(match45):
        coord = '0'
        #print(match45)
    #print("coord",coord)
    if all(char.isalpha() or char.isspace() or char == "°"  for char in coord):
        #print("c'est un string")
        coord = 0
    elif all(char.isdigit() for char in coord):
        #print("C'est une suite de chiffres")
        if not(match41):
            coord = 0

    coord = str(coord)

    if(match44):
        #print(match44)
        nb1 = match44.group(1)
        nb2 = match44.group(2)
        firstnb2 = nb2[0]
        #print("first nb2",firstnb2)
        #print("nb2 type",type(nb2))
        if(int(nb1) == 1 ):
            #print(firstnb2)
            if(int(firstnb2) == 0 or int(firstnb2) == 1):
               coord=nb1+nb2[0]+"."+nb2[1:]
            else:
               #("44",match44)
               coord='0'
        elif(int(nb1)==3):
             if(int(firstnb2) > 0 and int(firstnb2) < 8):
               coord=nb1+nb2[0]+"."+nb2[1:]

        elif(int(nb1) >= 2 and int(nb1) <= 6):
            nb12= '3'+str(nb1)
            coord = nb12+"."+nb2
            #print("concat avec 3",coord)
        elif(int(nb1) >= 7 and int (nb1)<=11):
            #print("coooord",coord)
            coord = str(nb1+"."+nb2)
        else:
            coord = nb1+"."+nb2
        

    if(match42):
        nb1 = match42.group(1)
        #print(nb1)
        nb2 = match42.group(2)
        #print("nb2 type",type(nb2))

        if(int(nb1) >= 2 and int(nb1) <= 6):
            nb12= '3'+str(nb1)
            coord = nb12+"."+nb2
            #print("concat avec 3",coord)

    if(match0):
        nb1 = match0.group(1)
        nb2 = match0.group(2)
        res = nb1[:2]+"."+nb1[2:]+nb2
        #print("res match0",res)
        coord = str(res)
    
    #print("coord ",coord)

    if(match46):
        nb1 = match46.group(1)
        nb2 = match46.group(2)
        nb3 = match46.group(3)
        nb4 = match46.group(4)
        nb5 = nb3+"."+nb4
        res = nb1+"_"+nb2+"_"+nb5
        coord = res
        #print(nb1 , nb2 , nb3 , nb5)
    if(match47):
        nb1 = match47[0]
        res = nb1
        coord = res
    if(match48):
        #print(match48[0])
        nb1 = match48.group(1)+"_"+match48.group(2)+"_"+match48.group(3)
        coord = nb1

    if(match43):
        nb1 = match43.group(1)
        nb2 = match43.group(2)
        nb3 = match43.group(3)
        if(nb1 == '0'):
            res = '1'+str(nb1)+"_"+nb2+"_"+nb3
            #print("res match0",res)
        else:
            res = nb1[:2]+"."+nb1[2:]+nb2
        coord = str(res)

    if match3:
        degrees = match3.group(1)
        minutes = match3.group(2)
        seconds = match3.group(3)
        coord = f"{degrees}_{minutes}_{seconds}"
    
    if match4:
        nb1 = match4.group(1)
        nb2 = match4.group(2)
        if(nb1==nb2):
            coord=nb1
    if (match5):
        deg = match5.group(1)
        min = match5.group(2)
        min = min[:2] + '.' + min[2:]
        sec = match5.group(3)
        coord = deg+"_"+min+"_"+sec
        #print("coord match",coord)
    if(match6):
        nb1 = match6.group(1)
        nb2 = match6.group(2)
        coord = nb1+"."+nb2
    if(match7):
        nb1 = match7.group(1)
        nb2 = match7.group(2)
        coord = nb1+"."+nb2
        #print("coord",coord)
    if(match8):
        nb1 = match8.group(1)
        nb2 = match8.group(2)
        nb2 = nb2[:2] + '.' + nb2[2:]
        nb3 = int(nb1)+float(nb2)/60
        coord = str(nb3)
        #print("coord",coord)
    if(match9):
        nb1 = match9.group(1)
        nb2 = match9.group(2)
        nb22 = nb2[:2]
        nb3 = match9.group(3)
        nb33 = nb2[2:] +'.' + nb3
        res = nb1+"_"+nb22+"_"+nb33
        #print("nb3",res)
        coord = res
    if(match10):
        nb1 = match10.group(1)
        nb2 = match10.group(2)
        nb3 = match10.group(3)
        res = nb1+"_"+nb2+"_"+nb3
        #print("res",res)
        coord = res
    if(match11):
        nb1 = match11.group(1)
        nb2 = match11.group(2)
        nb3 = match11.group(3)
        res = nb1+"_"+nb2+"_"+nb3
        #print("res",res)
        coord = res
    if(match12):
        nb1 = match12.group(1)
        coord = nb1

    if(match13):
        nb1 = match13.group(1)
        nb2 = match13.group(2)
        nb3 = match13.group(3)
        res = nb1+"_"+nb2+"_"+nb3
        #print("res",res)
        coord = res

    if(match14):
        nb1 = match14.group(1)
        nb2 = match14.group(2)
        nb3 = match14.group(3)
        res = nb1+"_"+nb2+"_"+nb3
        #print("res",res)
        coord = res
    if(match15):
        nb1 = match15.group(1)
        nb2 = match15.group(2)
        nb3 = match15.group(3)
        res = nb1+"_"+nb2+"_"+nb3
        #print("res1",res)
        coord = res
    if(match16):
        nb1 = match16.group(1)
        nb2 = match16.group(2)
        nb3 = match16.group(3)
        res = nb1+"_"+nb2+"_"+nb3
        #("res",res)
        coord = res
    if(match17):
        nb1 = match17.group(1)
        nb2 = match17.group(2)
        nb3 = match17.group(3)
        nb33 = nb3[:2] +'.' + nb3[2:]
        res = nb1+"_"+nb2+"_"+nb33
        #print("res",res)
        coord = res
    if(match18):
        nb1 = match18.group(1)
        nb2 = match18.group(2)
        nb3 = match18.group(3)
        res = nb1+"_"+nb2+"_"+nb3
        #print("res",res)
        coord = res
    if(match19):
        nb1 = match19.group(1)
        nb2 = match19.group(2)
        nb3 = match19.group(3)
        res = nb1+"_"+nb2+"_"+nb3
        #print("res",res)
        coord = res
    if(match20):
        nb1 = match20.group(1)
        nb2 = match20.group(2)
        nb3 = match20.group(3)
        res = nb1+"_"+nb2+"_"+nb3
        #print("res",res)
        coord = res
    if(match21):
        #print(match21)
        nb1 = match21.group(1)
        nb2 = match21.group(2)
        nb3 = match21.group(3)
        res=""
        nb1 = nb1.strip()
        nb1 = int(nb1)
        #print('nb1 ', nb1)
        if(nb1 > 41 and nb1 < 47):
            nb11 = nb1 - 10
            res = str(nb11)+"_"+nb2+"_"+nb3
            #print("res entre 41 et 47",res)
        elif(nb1 > 47 or ( nb1 > 13 and nb1 < 30 )):
            res = '0'
            #print("in 13 28")
        elif(nb1 > 2 and nb1 < 8):
            nb12 = '3'+ str(nb1)
            res = str(nb12)+"_"+nb2+"_"+nb3
            #print("concat avec 3",res)
        elif(nb1 == 0):
            nb13 = '1' + str(nb1)
            res = str(nb13)+"_"+nb2+"_"+nb3
            #print("degré = 0")
        else:
            res = str(nb1)+"_"+nb2+"_"+nb3
            #print("res",res)
        
        coord = res 
 

    if(match22):
        nb1 = match22.group(1)
        nb2 = match22.group(2)
        nb3 = match22.group(3)
        res = nb1+"_"+nb2+"_"+nb3
        #print("res",res)
        coord = res
    if(match23):
        nb1 = match23.group(1)
        nb2 = match23.group(2)
        nb3 = match23.group(3)
        nb4 = match23.group(4)
        nb5 = match23.group(5)
        nb33= nb3+"."+nb4+nb5
        res = nb1+"_"+nb2+"_"+nb33
        #print("res",res)
        coord = res
    if(match24):
        nb1 = match24.group(1)
        nb2 = match24.group(2)
        nb3 = match24.group(3)
        res = nb1+"_"+nb2+"_"+nb3
        #print("res",res)
        coord = res
    if(match25):
        nb1 = match25.group(1)
        nb2 = match25.group(2)
        nb3 = match25.group(3)
        res = nb1+"_"+nb2+"_"+nb3
        #print("res",res)
        coord = res
    if(match26):
        nb1 = match26.group(1)
        nb2 = match26.group(2)
        res = nb1+" "+nb2
        #print("res",res)
        coord = res
    if(match27):
        nb1 = match27.group(1)
        nb2 = match27.group(2)
        nb3 = match27.group(3)
        nb4 = match27.group(4)
        res = nb1+"."+nb2+"_"+nb3+"."+nb4
        #print("res",res)
        coord = res
    if(match28):
        nb1 = match28.group(1)
        nb2 = match28.group(2)
        nb3 = match28.group(3)
        res = nb1+"_"+nb2+"_"+nb3
        #print("res",res)
        coord = res
    if(match29):
        nb1 = match29.group(1)
        nb2 = match29.group(2)
        nb3 = float(nb1)+float(nb2)
        res = str(nb3)
        #print("res",res)
        coord = res
    if(match30):
        nb1 = match30.group(1)
        nb2 = match30.group(2)
        nb3 = match30.group(3)
        res = nb1+"_"+nb2+"_"+nb3
        #print("res",res)
        coord = res
    if(match31):
        nb1 = match31.group(1)
        nb2 = match31.group(2)
        res = nb1
        #print("res",res)
        coord = res
    if(match32):
        nb1 = match32.group(1)
        nb2 = match32.group(2)
        nb3 = match32.group(3)
        nb4 = match32.group(4)
        res = nb1+" "+nb2+" "+nb3+"."+nb4
        #print("res",res)
        coord = res
    if(match33):
        nb1 = match33.group(1)
        nb2 = match33.group(2)
        nb3 = match33.group(3)
        res = nb1+" "+nb2+" "+nb3
        #print("res",res)
        coord = res
    if(match34):
        nb1 = match34.group(1)
        nb2 = match34.group(2)
        nb3 = match34.group(3)
        nb4 = match34.group(4)
        res = nb1+" "+nb2+" "+nb3+"."+nb4
        #print("res",res)
        coord = res
    if(match35):
        nb1 = match35.group(1)
        nb2 = match35.group(2)
        nb3 = match35.group(3)
        res = nb1+" "+nb2+" "+nb3
        #print("res",res)
        coord = res
    if(match36):
        nb1 = match36.group(1)
        nb2 = match36.group(2)
        nb3 = match36.group(3)
        res = nb1+" "+nb2+" "+nb3
        #print("res",res)
        coord = res
    if(match37):
        nb1 = match37.group(1)
        nb2 = match37.group(2)
        res = nb1+"."+nb2
        #print("res",res)
        coord = res
    if(match38):
        nb1 = match38.group(2)
        res = nb1
        #print("res",res)
        coord = res
    if(match39):
        nb1 = match39.group(1)
        nb2 = match39.group(2)
        res = nb1+"."+nb2
        #print("res",res)
        coord = res
    if(match40):
        nb1 = match40.group(1)
        nb2 = match40.group(2)
        nb3 = match40.group(3)
        nb4 = match40.group(4)
        res = nb1+"."+nb2+" "+nb3+"."+nb4
        #print("res",res)
        coord = res
   
    if coord.startswith('°'):
        if coord.startswith('°'):
            coord = "10" + coord
    if 'B' in coord:
        coord = coord.replace('B', 'N', 1)  
        #print("coord avec B:", coord)
    if re.search(r'\s*E\w*.*$', coord):
        coord = re.sub(r'\s*E\w*.*$', 'E', coord)
    
    if re.search(r'\s*N\w*.*$', coord):
        coord = re.sub(r'\s*N\w*.*$', 'N', coord)

    if coord.count(',') == 2 and coord.endswith(','):
        coord = coord.replace(',', '.', 1)  
        #print("coord", coord)
        coord = re.sub(r"[\s°N,E',]", '_', coord)
    if match :
        coord = re.sub(r"[\s°N,E',]", '-', coord)
        #print("coord", coord)
    if(match2):
        parts = coord.split('.')
        #print("parts", parts)
        if len(parts) > 2:
            parts[2]='0.'+parts[2]
            coord = '.'.join(parts[:1] + [parts[1] + '-' + '.'.join(parts[2:])])
            #print("coords", coord)
    
    #print(coord)
    patternVerifNb=r"^(3\.[1-7]\d*\.?\d*|1\.[0-1]\d*|0\.[6-9]\d*)$"
    patternVerifNb1=r"^(0\.1[0-1]\d*|0\.3[1-7]\d*)$"
    patternVerisuite = r"^(\d+)\.(\d{1,2})$"



    matchVerif1 = re.search(patternVerifNb1, coord)
    matchVerifNb = re.search(patternVerifNb, coord)
    matchVerisuite = re.search(patternVerisuite, coord)
 
    if(matchVerisuite):
        nb1 = matchVerisuite.group(1)
        nb2 = matchVerisuite.group(2)
        res = nb1[:2] + '.' + nb1[2:] + nb2
        #print("res",res)
        coord = res

    if(matchVerif1):
        nb1 = matchVerif1.group(1)
        res = float(nb1)*100
        #print(res)
        coord = str(res)

    if(matchVerifNb):
        nb1 = matchVerifNb.group(1)
        res = float(nb1)*10
        #print(res)
        coord = str(res)
        
    if(match41):
        #print(match41)
        nb1 = match41.group(1)
        nb2 = match41.group(2)
        res = nb1+'.'+nb2
        #print("res 41",res)
        coord = str(res)

  
    else:
        coord = re.sub(r"[\s°N,E,S,F,R,Z',]", '_', coord)
    
    #print("coord finale",coord)
    return coord

# Fonction pour valider les coordonnées
def validate_coordinates(latitude, longitude):
    if not (6 <= longitude <= 11):
        longitude, latitude = latitude, longitude  
    if not (31 <= latitude <= 38):
        latitude, longitude = longitude, latitude  
    return latitude, longitude

def clean_crm_case(crm_case):
    timestamp_pattern = r"\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}" 
    matchTime= re.search(timestamp_pattern, str(crm_case))

    crm_case_str = str(crm_case)
    parts = crm_case_str.split(' ') 

    if('error_case_' in str(crm_case).lower())or(str(crm_case).lower() == 'error_case' )or(len(parts)==2):
        #print("crm case",crm_case)
        return None
    elif(str(crm_case).lower() == 'error_case'  ):
        print("crm case vide ",crm_case)
        return None
    elif(len(parts)==2):
        #print("lenth>2",crm_case)
        if(matchTime):
                #print("time",crm_case)
                return None  
    return(crm_case)

def check_underscores(word):
    #print("word",word)
    if (word.count('-') == 1 and not word.endswith('-')) or(word.count('-') == 2 and not word.endswith('--')) or(word.count('-') == 3 and word.endswith('--')):
        parts = word.split('-')
        parts = [part.strip() for part in parts if part.strip() != '']
        #print("parts",parts)
        if len(parts) == 2:
                try:
                      degrees = float(parts[0])
                      minutes = float(parts[1])
                      decimal = degrees + (minutes / 60)
                      return decimal,"none"
                except ValueError:
                      return word, "Format invalide pour DMS"
    if word.count('_') == 1 :
        #print("word avec _",word)
        if(word.startswith('_')):
            word = word.strip('_')
        parts = word.split('_')
        #print("parts _",parts)
        if len(parts) == 2  :
            latitude = parts[0]
            longitude = parts[1]
            return latitude, longitude 
        else:
            return word, None  
    elif word.count('_') == 2 and word.endswith('_') and word.startswith('_'):
        parts = word.strip('_')
        #print("parts",parts)
        return parts, None 
    elif word.count('_') == 2 and word.endswith('__'):
        parts = word.split('__')
        if len(parts) == 1 or len(parts) == 2:
            return parts[0], None  

    elif word.count('_') == 4 or(word.count('_') == 2 and not word.endswith('_')) or (word.count('_') == 3 and word.endswith('_') and word.count('__')==0):
        parts = word.split('_')
        parts = [part.strip() for part in parts if part.strip() != '']
        #print(parts)
        if len(parts)== 4:
            parts = parts[:-2] + ['.'.join(parts[-2:])]
        if len(parts) == 3:
            #print("parts",parts)
            try:
                degrees = float(parts[0])
                minutes = float(parts[1])
                seconds = float(parts[2])
                decimal = degrees + (minutes / 60) + (seconds / 3600)
                return decimal, "4 underscores"  
            except ValueError:
                return word, "Format invalide pour DMS"  
    elif word.count('_') == 5:
        parts = word.split('_')
        parts = [part.strip() for part in parts if part.strip() != '']
        #print("parts 5",parts)
        if len(parts) == 3:
            #print("parts",parts)
            try:
                degrees = float(parts[0])
                minutes = float(parts[1])
                seconds = float(parts[2])
                decimal = degrees + (minutes / 60) + (seconds / 3600)
                return decimal, "4 underscores"  
            except ValueError:
                return word, "Format invalide pour DMS" 
    return word, None  
   
cleaned_data = []
Erroned_data = []

for row in sheet.iter_rows(min_row=2):  
    crm_case = row[0].value  
    crm_case_cleaned = clean_crm_case(crm_case)
    DATE_CREATION_CRM = row[1].value
    MSISDN = row[6].value
    CONTACT_CLIENT = row[5].value
    CONTACT2_CLIENT = row[4].value

    client = row[26].value
    LATITUDE_SITE_init = row[9].value
    LONGITUDE_SITE_init = row[10].value
    client_nettoyé = client
    STT = row[14].value
    REP_TRAVAUX_STT = row[18].value
    DATE_FIN_TRV = row[19].value
    DATE_FIN_TRV = verif_date(DATE_FIN_TRV)
    dateAff_STT = row[13].value
    verifSTT,dateAff_STT,REP_TRAVAUX_STT =verif_affectation(STT,dateAff_STT,REP_TRAVAUX_STT)
    REP_TRAVAUX_STT = verif_rep_STT(REP_TRAVAUX_STT)
    delegation = row[30].value
    gouvernorat = row[8].value
    delegation_clean = verifierDelegation(delegation)
    gouvernorat_clean = verifierGouv(gouvernorat)
    verifSTT = verif_Nom_STT(STT)
    offre = row[28].value
    pack = row[29].value
    OPENING_DATE_SUR_TIMOS = row[2].value
    #print("opening date ", OPENING_DATE_SUR_TIMOS)

    STATUT = row[3].value
    REP_RDV = row[15].value
    DATE_PRISE_RDV = row[16].value
    CMT_RDV = row[17].value
    Detail = row[7].value
    Description = row[11].value
    DATE_PRISE_RDV, DATE_FIN_TRV = verif_dates(OPENING_DATE_SUR_TIMOS, DATE_PRISE_RDV, DATE_FIN_TRV)
    DATE_FIN_TRV = verif_date(DATE_FIN_TRV)
    DATE_PRISE_RDV= verif_date(DATE_PRISE_RDV)


    if len(row) > 2:
        latitude = clean_coordinate(row[9].value)
        longitude = clean_coordinate(row[10].value)
        #print("latitude cleaned",latitude)
        #print("longitude cleaned",longitude)
        
        if(latitude == '0' and longitude != '0'):
            part1, part2 = separate_numbers(longitude)

        else :
            part1, part2 = separate_numbers(latitude)
        #print("part1",part1,"part2",part2)


        if(part2 is not None):
             if (7 <= float(part1) <= 11):
                longitude = part1.split('_')
                latitude = part2
                if isinstance(longitude, list) and len(longitude) == 1 :
                    #print("longitude ",longitude[0])
                    #print("latitude ",latitude)
                    longitude=longitude[0]
             else:
                longitude = part2
                latitude = part1

        if((longitude == '_one' ) and latitude != '_one') :
            cleaned_latitude, latitude_result = check_underscores(latitude)
            if (7 <= float(cleaned_latitude) <= 11):
                longitude = cleaned_latitude
                latitude = latitude_result
            else:
                longitude = latitude_result
                latitude = cleaned_latitude
        cleaned_latitude, latitude_result = check_underscores(latitude)
        cleaned_longitude, longitude_result = check_underscores(longitude)
    else:  
        coordinate = clean_coordinate(row[9].value)
        cleaned_latitude, cleaned_longitude = check_underscores(coordinate)

    try:
        cleaned_latitude = float(cleaned_latitude)
        cleaned_longitude = float(cleaned_longitude)
    except (ValueError, TypeError):
        print(f"Erreur de conversion pour le cas CRM {crm_case}: latitude ou longitude invalide")
        continue

    if(crm_case_cleaned != None):
            try:
                  validated_latitude, validated_longitude = validate_coordinates(cleaned_latitude, cleaned_longitude)
                  cleaned_data.append({
                  'CRM_CASE': crm_case_cleaned,
                  'LATITUDE_SITE': validated_latitude,
                  'LONGITUDE_SITE': validated_longitude ,
                  'DATE_CREATION_CRM': DATE_CREATION_CRM,
                  'MSISDN': MSISDN,
                  'CONTACT_CLIENT': CONTACT_CLIENT,
                  'CONTACT2_CLIENT':CONTACT2_CLIENT,
                  'CLIENT': client_nettoyé,
                  'REP_TRAVAUX_STT': REP_TRAVAUX_STT,
                  'NAME_STT': STT,
                  'Delegation': delegation,
                  'Gouvernorat': gouvernorat,
                  'DATE_AFFECTATION_STT': dateAff_STT,
                  'DES_PACK': pack,
                  'offre': offre,
                  'OPENING_DATE_SUR_TIMOS': OPENING_DATE_SUR_TIMOS,
                  'STATUT': STATUT,
                  'REP_RDV': REP_RDV,
                  'DATE_PRISE_RDV': DATE_PRISE_RDV,
                  'CMT_RDV':CMT_RDV,
                   'Detail':Detail,
                   'DATE_FIN_TRV':DATE_FIN_TRV,
                   'Description':Description

                   })
            except Exception as e:
                   print(f"Erreur de validation pour le cas CRM {crm_case}: {e}")
    else:
        Erroned_data.append({
                  'CRM_CASE': crm_case,
                  'LATITUDE_SITE': LATITUDE_SITE_init,
                  'LONGITUDE_SITE': LONGITUDE_SITE_init ,
                  'DATE_CREATION_CRM': DATE_CREATION_CRM,
                  'MSISDN': MSISDN,
                  'CONTACT_CLIENT': CONTACT_CLIENT,
                  'CONTACT2_CLIENT':CONTACT2_CLIENT,
                  'CLIENT': client_nettoyé,
                  'REP_TRAVAUX_STT': REP_TRAVAUX_STT,
                  'NAME_STT': STT,
                  'Delegation': delegation,
                  'Gouvernorat': gouvernorat,
                  'DATE_AFFECTATION_STT': dateAff_STT,
                  'DES_PACK': pack,
                  'offre': offre,
                  'OPENING_DATE_SUR_TIMOS': OPENING_DATE_SUR_TIMOS,
                  'STATUT': STATUT,
                  'REP_RDV': REP_RDV,
                  'DATE_PRISE_RDV': DATE_PRISE_RDV,
                  'CMT_RDV':CMT_RDV,
                   'Detail':Detail,
                   'DATE_FIN_TRV':DATE_FIN_TRV,
                   'Description':Description

                   })
  
grouped_data = defaultdict(list)
for entry in cleaned_data:
    crm_case = entry['CRM_CASE']
    if crm_case:
        grouped_data[crm_case].append(entry)

filtered_cleaned_data = []

for crm_case, entries in grouped_data.items():
    sorted_entries = sorted(
        entries,
        key=lambda x: x['DATE_CREATION_CRM'] if isinstance(x['DATE_CREATION_CRM'], datetime) else datetime.min,
        reverse=True
    )
    filtered_cleaned_data.append(sorted_entries[0])

# Connexion à la base et insertion des données
connection = connect_to_db()
if connection and filtered_cleaned_data:
    insert_new_data(connection, filtered_cleaned_data)
    connection.close()
elif not filtered_cleaned_data:
    print("Aucune donnée valide à insérer.")

# Création des fichiers Excel
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active
new_sheet.title = "Cleaned Data"

new_workbookError = openpyxl.Workbook()
new_sheetError = new_workbookError.active
new_sheetError.title = "Erroned Data"

# Ajout des en-têtes
headers = ['CRM_CASE', 'DATE_CREATION_CRM', 'LATITUDE_SITE', 'LONGITUDE_SITE', 'MSISDN', 'CONTACT_CLIENT', 'CONTACT2_CLIENT','CLIENT','REP_TRAVAUX_STT','NAME_STT', 'Delegation', ' Gouvernorat' , 'DATE_AFFECTATION_STT ','DES_PACK','offre',
           'OPENING_DATE_SUR_TIMOS','REP_RDV','DATE_PRISE_RDV','CMT_RDV','Detail','STATUT','DATE_FIN_TRV','Description']
new_sheet.append(headers)
new_sheetError.append(headers)

# Remplissage des fichiers avec les données
for data in Erroned_data:
    new_sheetError.append([data['CRM_CASE'], data['DATE_CREATION_CRM'], data['LATITUDE_SITE'], 
                           data['LONGITUDE_SITE'], data['MSISDN'], data['CONTACT_CLIENT'],data['CONTACT2_CLIENT'], data['CLIENT'],data['REP_TRAVAUX_STT'], data['NAME_STT'] , data['Delegation']
                           ,data['Gouvernorat'],data['DATE_AFFECTATION_STT'],data['DES_PACK'],data['offre'],data['OPENING_DATE_SUR_TIMOS'],data['REP_RDV'],data['DATE_PRISE_RDV'],data['CMT_RDV']
                           ,data['Detail'],data['STATUT'],data['DATE_FIN_TRV'],data['Description']])


for data in cleaned_data:
    new_sheet.append([data['CRM_CASE'], data['DATE_CREATION_CRM'], data['LATITUDE_SITE'], 
                      data['LONGITUDE_SITE'], data['MSISDN'], data['CONTACT_CLIENT'],data['CONTACT2_CLIENT'], data['CLIENT'], data['REP_TRAVAUX_STT'],data['NAME_STT'], data['Delegation']
                           ,data['Gouvernorat'],data['DATE_AFFECTATION_STT'],data['DES_PACK'],data['offre'],data['OPENING_DATE_SUR_TIMOS'],data['REP_RDV'],data['DATE_PRISE_RDV']
                           ,data['CMT_RDV'],data['Detail'],data['STATUT'],data['DATE_FIN_TRV'],data['Description']])

# Sauvegarde des fichiers avec gestion des erreurs
try:
    new_file_pathError = './Erroned_dataPlainte.xlsx'
    new_workbookError.save(new_file_pathError)
    print(f"Les données erronées ont été enregistrées dans : {new_file_pathError}")
except Exception as e:
    print(f"Erreur lors de l'enregistrement des données erronées : {e}")

try:
    new_file_path = './cleaned_data_Plaintes.xlsx'
    new_workbook.save(new_file_path)
    print(f"Les données nettoyées ont été enregistrées dans : {new_file_path}")
except Exception as e:
    print(f"Erreur lors de l'enregistrement des données nettoyées : {e}")

cleaned_data = remove_duplicates(cleaned_data)


sql_file_path = './insert_plaintes.sql'

try:
    with open(sql_file_path, 'w', encoding='utf-8') as sql_file:
        sql_file.write("-- Script d'insertion pour la table plainte\n")
        sql_file.write("-- Généré automatiquement à partir des données nettoyées\n")
        sql_file.write(f"-- {len(cleaned_data)} enregistrements uniques (doublons supprimés)\n\n")
        
        for data in cleaned_data:
            insert_stmt = generate_sql_insert(data)
            sql_file.write(insert_stmt + "\n")
            
    print(f"Fichier SQL généré avec succès : {sql_file_path}")
    print(f"Nombre d'enregistrements uniques : {len(cleaned_data)}")
    
except Exception as e:
    print(f"Erreur lors de la génération du fichier SQL : {e}")
    
