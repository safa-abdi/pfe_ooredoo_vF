import datetime
import openpyxl
import re
from datetime import datetime, timedelta
import calendar
import mysql.connector
from mysql.connector import Error, errorcode

# Charger le fichier Excel existant
file_path = './data to clean/resiliation.xlsx'
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active



def connect_to_mysql():
    try:
        connection = mysql.connector.connect(
            host='localhost',       
            database='ooredoo',  
            user='root',            
            password='#Safa@123_#'
        )
        if connection.is_connected():
            cursor = connection.cursor()
            create_table(cursor)
            return connection, cursor
    except Error as err:
        if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
            print("❌ Échec de connexion : Mauvais login/mot de passe")
        elif err.errno == errorcode.ER_BAD_DB_ERROR:
            print("❌ La base de données n'existe pas")
        else:
            print(err)
        return None, None

def create_table(cursor):
    create_query = """
    CREATE TABLE IF NOT EXISTS resiliation (
        CRM_CASE INT AUTO_INCREMENT PRIMARY KEY,
        DATE_CREATION_CRM DATETIME,
        LATITUDE_SITE DOUBLE NULL DEFAULT NULL,
        LONGITUDE_SITE DOUBLE NULL DEFAULT NULL,
        MSISDN VARCHAR(20),
        CONTACT1_CLIENT VARCHAR(100),
        CONTACT2_CLIENT VARCHAR(100),
        CLIENT TEXT,
        REP_TRAVAUX_STT TEXT,
        STT TEXT,
        Delegation TEXT,
        Gouvernorat TEXT,
        DATE_AFFECTATION_STT DATETIME NULL DEFAULT NULL,
        DES_PACK TEXT,
        OPENING_DATE_SUR_TIMOS DATETIME,
        STATUT TEXT,
        REP_RDV TEXT,
        DATE_PRISE_RDV DATETIME NULL DEFAULT NULL,
        CMT_RDV TEXT,
        DATE_FIN_TRV DATETIME NULL DEFAULT NULL,
        Description TEXT,
        Detail TEXT,
        entite TEXT
    );
    """
    cursor.execute(create_query)

def remove_duplicates(data_list):
    unique_records = {}
    
    for data in data_list:
        identifier = data['CRM_CASE']
        current_date = data.get('DATE_CREATION_CRM')
        if identifier not in unique_records or (
            current_date and unique_records[identifier].get('DATE_CREATION_CRM') and 
            current_date > unique_records[identifier].get('DATE_CREATION_CRM')
        ):
            unique_records[identifier] = data
    
    return list(unique_records.values())
def parse_date(date_val):
    if date_val in [None, "NULL", "", "null"]:
        return None
    if isinstance(date_val, datetime):
        return date_val.strftime("%Y-%m-%d %H:%M:%S")
    try:
        for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y"]:
            try:
                dt = datetime.strptime(str(date_val), fmt)
                return dt.strftime("%Y-%m-%d %H:%M:%S")
            except ValueError:
                continue
        print(f"⚠️ Date non reconnue : {date_val}")
        return None
    except Exception as e:
        print(f"❌ Erreur lors du parsing de la date : {date_val} | {e}")
        return None
    
def verif_rep_STT(rep):
    return rep if rep else "en_cours" 

def verif_affectation(nomSTT, dateAff, REP_TRAVAUX_STT):
    #print("stt",nomSTT)
    if not nomSTT and not dateAff and not REP_TRAVAUX_STT:
        return "non_affecté", None, "non_affecté_stt"
    return nomSTT, dateAff, REP_TRAVAUX_STT
    
def safe_float(val):
    try:
        if val is None:
            return None
        val = str(val).strip()
        if val.lower() in ('none', 'null', '', 'nan'):
            return None
        f_val = float(val)
        if -90 <= f_val <= 90:
            return f_val
        elif -180 <= f_val <= 180:
            return f_val
        else:
            print(f"⚠️ Valeur hors plage géographique : {f_val}")
            return None
    except (ValueError, TypeError):
        print(f"❌ Impossible de convertir en float: {val}")
        return None
    
def verifierDelegation(deleg):
    if deleg is None:
        return None
    return None if "@" in deleg else deleg
def verifierGouv(gouv):
    if gouv is None:
        return None
    return gouv if gouv.isalpha() else "NULL"


def verif_date(dateRDV):
    if not dateRDV or dateRDV in ["NULL", "None", ""]:
        return  None
    if isinstance(dateRDV, datetime):
        parsed_date = dateRDV
    else:
        formats = ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y"]
        parsed_date = None
        for fmt in formats:
            try:
                parsed_date = datetime.strptime(str(dateRDV), fmt)
                break
            except ValueError:
                continue

        if not parsed_date:
            print(f"⚠️ Format de date invalide : {dateRDV}")
            return "NULL"
    if parsed_date.hour == 0 and parsed_date.minute == 0 and parsed_date.second == 0:
        parsed_date = parsed_date.replace(hour=23, minute=59, second=0)

    return parsed_date.strftime("%Y-%m-%d %H:%M:%S")
def verif_cable(cable):
    if not cable: 
        return "NULL"
    match = re.search(r'(\d+)', str(cable)) 
    return int(match.group(1)) if match else "NULL" 

def separate_numbers(input_str):
    match = re.match(r"(\d+\.\d+)\.(\d+\.\d+)", input_str)
    match100 = re.match(r"(\d+\.\d+)\_(\d+\.\d+)", input_str)
    match200 = re.match(r"(\d+\.\d+)", input_str) 
    if match200:
        return match200.group(1),None 
    if match :
        return match.group(1), match.group(2)  
    if match100 :
        return match100.group(1), match100.group(2)  
    else:
        return input_str, None  

from datetime import datetime, timedelta


def verif_dates(OPENING_DATE_SUR_TIMOS, DATE_PRISE_RDV, DATE_FIN_TRV):
    def parse_date(date_val):
        if date_val in [None, "NULL", "", "null"]:
            return None
        if isinstance(date_val, str):
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                try:
                    return datetime.strptime(date_val, fmt)
                except ValueError:
                    continue
        return date_val if isinstance(date_val, datetime) else None

    opening_date = parse_date(OPENING_DATE_SUR_TIMOS)
    date_prise_rdv = parse_date(DATE_PRISE_RDV)
    date_fin_trv = parse_date(DATE_FIN_TRV)
    sysdate = datetime.now()

    def adjust_future_date(date_to_adjust, reference_date):
        if reference_date.date() == sysdate.date():  # Même jour
            return sysdate + timedelta(hours=2)
        else:
            return reference_date + timedelta(days=2)

    def adjust_date(date_to_check, reference_date):
        if not date_to_check or not reference_date:
            return None

        # Vérification futur
        if date_to_check > sysdate:
            return adjust_future_date(date_to_check, reference_date)

        year = date_to_check.year
        month = date_to_check.month
        day = date_to_check.day

        if year < reference_date.year:
            year = reference_date.year

        if year == reference_date.year:
            if month < reference_date.month:
                month = reference_date.month
            elif month == reference_date.month and day < reference_date.day:
                date_to_check += timedelta(days=2)
                return date_to_check if date_to_check <= sysdate else adjust_future_date(date_to_check, reference_date)

        # Vérifie que le jour est valide pour le mois
        last_day_of_month = calendar.monthrange(year, month)[1]
        if day > last_day_of_month:
            day = last_day_of_month

        try:
            date_to_check = date_to_check.replace(year=year, month=month, day=day)
        except ValueError:
            date_to_check = date_to_check.replace(year=year, month=month, day=last_day_of_month)

        return date_to_check if date_to_check <= sysdate else adjust_future_date(date_to_check, reference_date)

    if opening_date:
        date_prise_rdv = adjust_date(date_prise_rdv, opening_date)
        date_fin_trv = adjust_date(date_fin_trv, opening_date)

    return date_prise_rdv, date_fin_trv
def clean_coordinate(coord):
    coord = str(coord)
    match = re.match(r"(\d+)°(\d+\.\d+)\s*([NSEWZ])", coord)
    match2 = re.match(r'^\d+\.\d{5,}\.\d{2}$', coord)
    pattern = r"(\d{1,3})\.(\d{1,3})°(\d{1,3}\.\d+)\s*[NSEWZ]"
    pattern0 = r"^(\d+)\.(\d+)°[NSEWZ]$"
    pattern00 = r"^(\d+)\.(\d+)°$"
    pattern2 = r"(\d{1,3}\.\d+)[°]?(\d{1,3}\.\d+)"
    pattern3 = r"^(\d+)\.(\d+)\.(\d+)$"
    pattern4 = r"^(\d+)\.\.(\d+)$"
    pattern5 = r"(\d+),(\d+)°([NSEZ])"
    pattern6 = r"^(\d+)°(\d+) ([NSEZ])$"
    pattern7 = r"^(\d+)°(\d+)\.(\d+) ([NSEZ])$"
    pattern8 = r"^LAT(\d+)°(\d+) (\d+\.\d+) ([NSEZ])$"
    pattern9 = r"^LO°N(\d+) (\d+) (\d+\.\d+) ([NSEZ])$"
    pattern10 = r"^(\d+\.\d+)\.$"
    pattern11 = r"^(\d+)\.(\d+)°(\d+) ([NSEZ])$"
    pattern12 = r"^(\d+)\.(\d+)\.(\d+)°([NSEZ])$"
    pattern13 = r"(\d+)°(\d+)'(\d+\.\d+)$"
    pattern14 = r"(\d+)°(\d+)'(\d+\.\d+)([NSWE])"
    pattern15 = r"^(\d+)°(\d+)\.(\d+)'([NSWE])$"
    pattern16 = r"(\d+)\)(\d+)°(\d+\.\d+) ([NSWE])"
    pattern17 = r"(\d+)°(\d+) \.(\d+\.\d+) ([NSWE])"
    pattern18 = r"(\d+)°\.(\d+) (\d+\.\d+) ([NSWE])"
    pattern19 = r"(\d+)°(\d+) (\d+\.\d+) ([NSWE])"
    pattern20 = r"(\d+)° (\d+)'(\d+\.\d+)\"([NSWE])"
    pattern21 = r"^(\d+)°(\d+) (\d+) (\d+)\.(\d+) ([NSWE])$"
    pattern22 = r"(\d+)°(\d+) (\d+\.\d+)\. ([NSWE])"
    pattern23 = r"(\d+)° (\d+)'(\d+\.\d+)$"
    pattern24 = r"(\d+\.\d+),(\d+\.\d+)"
    pattern25 = r"(\d+)\,(\d+)\/(\d+)\,(\d+)"
    pattern26 = r"(\d{1,2})°(\d{1,2}) (\d{1,2}\.\d+)\* ([NSEW])"
    pattern27 = r"(\d+\.\d+)\+(\d+)"
    pattern28 = r"(\d+)°(\d+)'(\d+\.\d+)\" ([NSWE])"
    pattern29 = r"(\d+\.\d+)°\s*(\d+)'?"
    pattern30 = r"^(\d+)°(\d+)\.(\d+)\.(\d+) ([NSEZ])$"
    pattern31 = r"^(\d+)°(\d+) \.(\d+)\. ([NSEZ])$"
    pattern32 = r"^(\d+)°(\d+) (\d+) \.(\d+) ([NSEZ])$"
    pattern33 = r"^(\d+)\.(\d+)\.(\d+)\.°([NSEZ])$"
    pattern34 = r"^(\d+)\.(\d+)\.(\d+)([NSEZF])$"
    pattern35 = r"^(\d+)\,(\d+)°$"
    pattern36 =r"^([A-Za-z]+)(\d+\.\d+)"
    pattern37 = r"^(\d+)\.(\d+)\.(\d+)°$"
    pattern38 = r"^(\d+)\.(\d+)\.(\d+)\.(\d+)$"
    pattern39 = r"^(3[1-7]|1[01])(\d+)$"
    pattern40 = r"^(\d)\.(\d+)°$"
    pattern43 = r"^(\d+)°(\d+) (\d+\.\d+) $"
    pattern44 = r"^(\d+)\.(\d+)$"
    pattern45 = r"^(\d+)\.(\d+)\.(\d+)\.(\d+)\.(\d+)$"
    pattern50 = r"^(\d+)°\.(\d+) \.(\d+)\.(\d+) ([NSEZF])$"
    match00 =re.search(pattern00,coord)
    match45 = re.search(pattern45, coord)
    match44 = re.search(pattern44, coord)
    match43 = re.search(pattern43, coord)
    # Vérifier le match
    match3 = re.search(pattern, coord)
    match4 = re.search(pattern2, coord)
    match5 = re.search(pattern3, coord)
    match6 = re.search(pattern4, coord)
    match7 = re.search(pattern5, coord)
    match8 = re.search(pattern6, coord)
    match9 = re.search(pattern7, coord)
    match10 = re.search(pattern8, coord)
    match11 = re.search(pattern9, coord)
    match12 = re.search(pattern10, coord)
    match13 = re.search(pattern11, coord)
    match14 = re.search(pattern12, coord)
    match15 = re.search(pattern13, coord)
    match16 = re.search(pattern14, coord)
    match17 = re.search(pattern15, coord)
    match18 = re.search(pattern16, coord)
    match19 = re.search(pattern17, coord)
    match20 = re.search(pattern18, coord)
    match21 = re.search(pattern19, coord)
    match22 = re.search(pattern20, coord)
    match23 = re.search(pattern21, coord)
    match24 = re.search(pattern22, coord)
    match25 = re.search(pattern23, coord)
    match26 = re.search(pattern24, coord)
    match27 = re.search(pattern25, coord)
    match28 = re.search(pattern26, coord)
    match29 = re.search(pattern27, coord)
    match30 = re.search(pattern28, coord)
    match31 = re.search(pattern29, coord)
    match32 = re.search(pattern30, coord)
    match33 = re.search(pattern31, coord)
    match34 = re.search(pattern32, coord)
    match35 = re.search(pattern33, coord)
    match36 = re.search(pattern34, coord)
    match37 = re.search(pattern35, coord)
    match38 = re.search(pattern36, coord)
    match39 = re.search(pattern37, coord)
    match40 = re.search(pattern38, coord)
    match41 = re.match(pattern39, str(coord))
    match0 = re.match(pattern0, str(coord))
    match42 = re.match(pattern40, str(coord))
    match50 = re.match(pattern50, str(coord))

    if(match00):
        nb1 = match00.group(1)
        nb2 = match00.group(2)
        nb20= int(nb2[0])
        nb21= int(nb2[1])
        if(int(nb1)==0):
            if(nb20 == 1 and (nb21 == 0 or nb21 ==1)):
                coord=nb2[0]+nb2[1]+'.'+nb2[2:]
            elif(nb20 == 3 and (nb21 > 0 and nb21 < 8)):
                coord=nb2[0]+nb2[1]+'.'+nb2[2:]
        #print("coord",coord)

    if(match45):
        coord = '0'
        #print(match45)
    #print("coord",coord)
    if all(char.isalpha() or char.isspace() or char == "°"  for char in coord):
        #print("c'est un string")
        coord = 0
    elif all(char.isdigit() for char in coord):
        #print("C'est une suite de chiffres")
        if not(match41):
            coord = 0

    coord = str(coord)

    if(match44):
        #print(match44)
        nb1 = match44.group(1)
        nb2 = match44.group(2)
        firstnb2 = nb2[0]
        #print("first nb2",firstnb2)
        #print("nb2 type",type(nb2))
        if(int(nb1) == 1 ):
            #print(firstnb2)
            if(int(firstnb2) == 0 or int(firstnb2) == 1):
               coord=nb1+nb2[0]+"."+nb2[1:]
            else:
               #("44",match44)
               coord='0'
        elif(int(nb1)==3):
             if(int(firstnb2) > 0 and int(firstnb2) < 8):
               coord=nb1+nb2[0]+"."+nb2[1:]

        elif(int(nb1) >= 2 and int(nb1) <= 6):
            nb12= '3'+str(nb1)
            coord = nb12+"."+nb2
            #print("concat avec 3",coord)
        elif(int(nb1) >= 7 and int (nb1)<=11):
            #print("coooord",coord)
            coord = str(nb1+"."+nb2)
        else:
            coord = nb1+"."+nb2
        

    if(match42):
        nb1 = match42.group(1)
        #print(nb1)
        nb2 = match42.group(2)
        #print("nb2 type",type(nb2))

        if(int(nb1) >= 2 and int(nb1) <= 6):
            nb12= '3'+str(nb1)
            coord = nb12+"."+nb2
            #print("concat avec 3",coord)

    if(match0):
        nb1 = match0.group(1)
        nb2 = match0.group(2)
        res = nb1[:2]+"."+nb1[2:]+nb2
        #print("res match0",res)
        coord = str(res)
    if(match50):
        nb1 = match50.group(1)
        nb2 = match50.group(2)
        nb3 = match50.group(3)+"."+match50.group(4)
        res = nb1+"_"+nb2+"_"+nb3
        #print("res match50",coord)
        coord = str(res)

    if(match43):
        nb1 = match43.group(1)
        nb2 = match43.group(2)
        nb3 = match43.group(3)
        if(nb1 == '0'):
            res = '1'+str(nb1)+"_"+nb2+"_"+nb3
            #print("res match0",res)
        else:
            res = nb1[:2]+"."+nb1[2:]+nb2
        coord = str(res)

    if match3:
        degrees = match3.group(1)
        minutes = match3.group(2)
        seconds = match3.group(3)
        coord = f"{degrees}_{minutes}_{seconds}"
    
    if match4:
        nb1 = match4.group(1)
        nb2 = match4.group(2)
        if(nb1==nb2):
            coord=nb1
    if (match5):
        deg = match5.group(1)
        min = match5.group(2)
        min = min[:2] + '.' + min[2:]
        sec = match5.group(3)
        coord = deg+"_"+min+"_"+sec
        #print("coord match",coord)
    if(match6):
        nb1 = match6.group(1)
        nb2 = match6.group(2)
        coord = nb1+"."+nb2
    if(match7):
        nb1 = match7.group(1)
        nb2 = match7.group(2)
        coord = nb1+"."+nb2
        #print("coord",coord)
    if(match8):
        nb1 = match8.group(1)
        nb2 = match8.group(2)
        nb2 = nb2[:2] + '.' + nb2[2:]
        nb3 = int(nb1)+float(nb2)/60
        coord = str(nb3)
        #print("coord",coord)
    if(match9):
        nb1 = match9.group(1)
        nb2 = match9.group(2)
        nb22 = nb2[:2]
        nb3 = match9.group(3)
        nb33 = nb2[2:] +'.' + nb3
        res = nb1+"_"+nb22+"_"+nb33
        #print("nb3",res)
        coord = res
    if(match10):
        nb1 = match10.group(1)
        nb2 = match10.group(2)
        nb3 = match10.group(3)
        res = nb1+"_"+nb2+"_"+nb3
        #print("res",res)
        coord = res
    if(match11):
        nb1 = match11.group(1)
        nb2 = match11.group(2)
        nb3 = match11.group(3)
        res = nb1+"_"+nb2+"_"+nb3
        #print("res",res)
        coord = res
    if(match12):
        nb1 = match12.group(1)
        coord = nb1

    if(match13):
        nb1 = match13.group(1)
        nb2 = match13.group(2)
        nb3 = match13.group(3)
        res = nb1+"_"+nb2+"_"+nb3
        #print("res",res)
        coord = res

    if(match14):
        nb1 = match14.group(1)
        nb2 = match14.group(2)
        nb3 = match14.group(3)
        res = nb1+"_"+nb2+"_"+nb3
        #print("res",res)
        coord = res
    if(match15):
        nb1 = match15.group(1)
        nb2 = match15.group(2)
        nb3 = match15.group(3)
        res = nb1+"_"+nb2+"_"+nb3
        #print("res1",res)
        coord = res
    if(match16):
        nb1 = match16.group(1)
        nb2 = match16.group(2)
        nb3 = match16.group(3)
        res = nb1+"_"+nb2+"_"+nb3
        #("res",res)
        coord = res
    if(match17):
        nb1 = match17.group(1)
        nb2 = match17.group(2)
        nb3 = match17.group(3)
        nb33 = nb3[:2] +'.' + nb3[2:]
        res = nb1+"_"+nb2+"_"+nb33
        #print("res",res)
        coord = res
    if(match18):
        nb1 = match18.group(1)
        nb2 = match18.group(2)
        nb3 = match18.group(3)
        res = nb1+"_"+nb2+"_"+nb3
        #print("res",res)
        coord = res
    if(match19):
        nb1 = match19.group(1)
        nb2 = match19.group(2)
        nb3 = match19.group(3)
        res = nb1+"_"+nb2+"_"+nb3
        #print("res",res)
        coord = res
    if(match20):
        nb1 = match20.group(1)
        nb2 = match20.group(2)
        nb3 = match20.group(3)
        res = nb1+"_"+nb2+"_"+nb3
        #print("res",res)
        coord = res
    if(match21):
        #print(match21)
        nb1 = match21.group(1)
        nb2 = match21.group(2)
        nb3 = match21.group(3)
        res=""
        nb1 = nb1.strip()
        nb1 = int(nb1)
        #print('nb1 ', nb1)
        if(nb1 > 41 and nb1 < 47):
            nb11 = nb1 - 10
            res = str(nb11)+"_"+nb2+"_"+nb3
            #print("res entre 41 et 47",res)
        elif(nb1 > 47 or ( nb1 > 13 and nb1 < 30 )):
            res = '0'
            #print("in 13 28")
        elif(nb1 > 2 and nb1 < 8):
            nb12 = '3'+ str(nb1)
            res = str(nb12)+"_"+nb2+"_"+nb3
            #print("concat avec 3",res)
        elif(nb1 == 0):
            nb13 = '1' + str(nb1)
            res = str(nb13)+"_"+nb2+"_"+nb3
            #print("degré = 0")
        else:
            res = str(nb1)+"_"+nb2+"_"+nb3
            #print("res",res)
        
        coord = res 
 

    if(match22):
        nb1 = match22.group(1)
        nb2 = match22.group(2)
        nb3 = match22.group(3)
        res = nb1+"_"+nb2+"_"+nb3
        #print("res",res)
        coord = res
    if(match23):
        nb1 = match23.group(1)
        nb2 = match23.group(2)
        nb3 = match23.group(3)
        nb4 = match23.group(4)
        nb5 = match23.group(5)
        nb33= nb3+"."+nb4+nb5
        res = nb1+"_"+nb2+"_"+nb33
        #print("res",res)
        coord = res
    if(match24):
        nb1 = match24.group(1)
        nb2 = match24.group(2)
        nb3 = match24.group(3)
        res = nb1+"_"+nb2+"_"+nb3
        #print("res",res)
        coord = res
    if(match25):
        nb1 = match25.group(1)
        nb2 = match25.group(2)
        nb3 = match25.group(3)
        res = nb1+"_"+nb2+"_"+nb3
        #print("res",res)
        coord = res
    if(match26):
        nb1 = match26.group(1)
        nb2 = match26.group(2)
        res = nb1+" "+nb2
        #print("res",res)
        coord = res
    if(match27):
        nb1 = match27.group(1)
        nb2 = match27.group(2)
        nb3 = match27.group(3)
        nb4 = match27.group(4)
        res = nb1+"."+nb2+"_"+nb3+"."+nb4
        #print("res",res)
        coord = res
    if(match28):
        nb1 = match28.group(1)
        nb2 = match28.group(2)
        nb3 = match28.group(3)
        res = nb1+"_"+nb2+"_"+nb3
        #print("res",res)
        coord = res
    if(match29):
        nb1 = match29.group(1)
        nb2 = match29.group(2)
        nb3 = float(nb1)+float(nb2)
        res = str(nb3)
        #print("res",res)
        coord = res
    if(match30):
        nb1 = match30.group(1)
        nb2 = match30.group(2)
        nb3 = match30.group(3)
        res = nb1+"_"+nb2+"_"+nb3
        #print("res",res)
        coord = res
    if(match31):
        nb1 = match31.group(1)
        nb2 = match31.group(2)
        res = nb1
        #print("res",res)
        coord = res
    if(match32):
        nb1 = match32.group(1)
        nb2 = match32.group(2)
        nb3 = match32.group(3)
        nb4 = match32.group(4)
        res = nb1+" "+nb2+" "+nb3+"."+nb4
        #print("res",res)
        coord = res
    if(match33):
        nb1 = match33.group(1)
        nb2 = match33.group(2)
        nb3 = match33.group(3)
        res = nb1+" "+nb2+" "+nb3
        #print("res",res)
        coord = res
    if(match34):
        nb1 = match34.group(1)
        nb2 = match34.group(2)
        nb3 = match34.group(3)
        nb4 = match34.group(4)
        res = nb1+" "+nb2+" "+nb3+"."+nb4
        #print("res",res)
        coord = res
    if(match35):
        nb1 = match35.group(1)
        nb2 = match35.group(2)
        nb3 = match35.group(3)
        res = nb1+" "+nb2+" "+nb3
        #print("res",res)
        coord = res
    if(match36):
        nb1 = match36.group(1)
        nb2 = match36.group(2)
        nb3 = match36.group(3)
        res = nb1+" "+nb2+" "+nb3
        #print("res",res)
        coord = res
    if(match37):
        nb1 = match37.group(1)
        nb2 = match37.group(2)
        res = nb1+"."+nb2
        #print("res",res)
        coord = res
    if(match38):
        nb1 = match38.group(2)
        res = nb1
        #print("res",res)
        coord = res
    if(match39):
        nb1 = match39.group(1)
        nb2 = match39.group(2)
        res = nb1+"."+nb2
        #print("res",res)
        coord = res
    if(match40):
        nb1 = match40.group(1)
        nb2 = match40.group(2)
        nb3 = match40.group(3)
        nb4 = match40.group(4)
        res = nb1+"."+nb2+" "+nb3+"."+nb4
        #print("res",res)
        coord = res
   
    if coord.startswith('°'):
        if coord.startswith('°'):
            coord = "10" + coord
    if 'B' in coord:
        coord = coord.replace('B', 'N', 1)  
        #print("coord avec B:", coord)
    if re.search(r'\s*E\w*.*$', coord):
        coord = re.sub(r'\s*E\w*.*$', 'E', coord)
    
    if re.search(r'\s*N\w*.*$', coord):
        coord = re.sub(r'\s*N\w*.*$', 'N', coord)

    if coord.count(',') == 2 and coord.endswith(','):
        coord = coord.replace(',', '.', 1)  
        #print("coord", coord)
        coord = re.sub(r"[\s°N,E',]", '_', coord)
    if match :
        coord = re.sub(r"[\s°N,E',]", '-', coord)
        #print("coord", coord)
    if(match2):
        parts = coord.split('.')
        #print("parts", parts)
        if len(parts) > 2:
            parts[2]='0.'+parts[2]
            coord = '.'.join(parts[:1] + [parts[1] + '-' + '.'.join(parts[2:])])
            #print("coords", coord)
    
    #print(coord)
    patternVerifNb=r"^(3\.[1-7]\d*\.?\d*|1\.[0-1]\d*|0\.[6-9]\d*)$"
    patternVerifNb1=r"^(0\.1[0-1]\d*|0\.3[1-7]\d*)$"
    patternVerisuite = r"^(\d+)\.(\d{1,2})$"



    matchVerif1 = re.search(patternVerifNb1, coord)
    matchVerifNb = re.search(patternVerifNb, coord)
    matchVerisuite = re.search(patternVerisuite, coord)
 
    if(matchVerisuite):
        nb1 = matchVerisuite.group(1)
        nb2 = matchVerisuite.group(2)
        res = nb1[:2] + '.' + nb1[2:] + nb2
        #print("res",res)
        coord = res

    if(matchVerif1):
        nb1 = matchVerif1.group(1)
        res = float(nb1)*100
        #print(res)
        coord = str(res)

    if(matchVerifNb):
        nb1 = matchVerifNb.group(1)
        res = float(nb1)*10
        #print(res)
        coord = str(res)
        
    if(match41):
        #print(match41)
        nb1 = match41.group(1)
        nb2 = match41.group(2)
        res = nb1+'.'+nb2
        #print("res 41",res)
        coord = str(res)

  
    else:
        coord = re.sub(r"[\s°N,E,S,F,R,Z',]", '_', coord)
    
    #print("coord finale",coord)
    return coord

def validate_coordinates(latitude, longitude):
    if not (6 <= longitude <= 11):
        longitude, latitude = latitude, longitude  
    if not (31 <= latitude <= 38):
        latitude, longitude = longitude, latitude  
    return latitude, longitude

def clean_crm_case(crm_case):
    crm_case_str = str(crm_case)
    parts = crm_case_str.split(' ') 

    if('ERROR_CASE_' in str(crm_case))or(str(crm_case) == 'ERROR_CASE' )or(len(parts)==2):
        return None
    return(crm_case)

def verif_Nom_STT(stt):
    timestamp_pattern = r"\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}" 
    matchTime= re.search(timestamp_pattern, str(crm_case))
    STT = str(stt)
    parts = STT.split(' ') 
    if (len(parts)==2):
        if(matchTime):
                return None 
    return stt

def check_underscores(word):
    if (word.count('-') == 1 and not word.endswith('-')) or(word.count('-') == 2 and not word.endswith('--')) or(word.count('-') == 3 and word.endswith('--')):
        parts = word.split('-')
        parts = [part.strip() for part in parts if part.strip() != '']
        if len(parts) == 2:
                try:
                      degrees = float(parts[0])
                      minutes = float(parts[1])
                      decimal = degrees + (minutes / 60)
                      return decimal,"none"
                except ValueError:
                      return word, "Format invalide pour DMS"
    if word.count('_') == 1 :
        if(word.startswith('_')):
            word = word.strip('_')
        parts = word.split('_')
        if len(parts) == 2  :
            latitude = parts[0]
            longitude = parts[1]
            return latitude, longitude 
        else:
            return word, None  
    elif word.count('_') == 2 and word.endswith('_') and word.startswith('_'):
        parts = word.strip('_')
        return parts, None 
    elif word.count('_') == 2 and word.endswith('__'):
        parts = word.split('__')
        if len(parts) == 1 or len(parts) == 2:
            return parts[0], None  

    elif word.count('_') == 4 or(word.count('_') == 2 and not word.endswith('_')) or (word.count('_') == 3 and word.endswith('_') and word.count('__')==0):
        parts = word.split('_')
        parts = [part.strip() for part in parts if part.strip() != '']
        #print(parts)
        if len(parts)== 4:
            parts = parts[:-2] + ['.'.join(parts[-2:])]
        if len(parts) == 3:
            #print("parts",parts)
            try:
                degrees = float(parts[0])
                minutes = float(parts[1])
                seconds = float(parts[2])
                decimal = degrees + (minutes / 60) + (seconds / 3600)
                return decimal, "4 underscores"  
            except ValueError:
                return word, "Format invalide pour DMS"  
    elif word.count('_') == 5:
        parts = word.split('_')
        parts = [part.strip() for part in parts if part.strip() != '']
        #print("parts 5",parts)
        if len(parts) == 3:
            #print("parts",parts)
            try:
                degrees = float(parts[0])
                minutes = float(parts[1])
                seconds = float(parts[2])
                decimal = degrees + (minutes / 60) + (seconds / 3600)
                return decimal, "4 underscores"  
            except ValueError:
                return word, "Format invalide pour DMS" 
    return word, None  
   

def nettoyer_NomClient(texte):
    texte = texte.replace("résiliation fixe JDID", "").strip()
    texte = re.sub(r'\b\d+\b$', '', texte).strip()
    return texte

cleaned_data = []
Erroned_data = []

for row in sheet.iter_rows(min_row=2):  
    crm_case = row[0].value  
    crm_case_cleaned = clean_crm_case(crm_case)
    DATE_CREATION_CRM = row[1].value
    MSISDN = row[4].value
    CONTACT1_CLIENT = row[3].value
    CONTACT2_CLIENT = row[10].value
    client = row[9].value
    description = row[8].value
    LATITUDE_SITE_init = row[5].value
    LONGITUDE_SITE_init = row[6].value
    client_nettoyé = nettoyer_NomClient(client)
    REP_TRAVAUX_STT = row[20].value if len(row) > 20 else None
    DATE_FIN_TRV = row[19].value if len(row) > 19 else None 
    DATE_FIN_TRV = verif_date(DATE_FIN_TRV)
    STT = row[14].value if len(row) > 14 else None 
    dateAff_STT = row[13].value if len(row) > 13 else None 
    verifSTT,dateAff_STT,REP_TRAVAUX_STT =verif_affectation(STT,dateAff_STT,REP_TRAVAUX_STT)
    REP_TRAVAUX_STT = verif_rep_STT(REP_TRAVAUX_STT)
    Detail = row[11].value
    entite = row[12].value
    delegation = row[28].value if len(row) > 28 else None 
    gouvernorat = row[7].value if len(row) > 7 else None 
    delegation_clean = verifierDelegation(delegation)
    gouvernorat_clean = verifierGouv(gouvernorat)
    verifSTT = verif_Nom_STT(STT)
    pack = row[29].value if len(row) > 14 else None 
    OPENING_DATE_SUR_TIMOS = row[2].value if len(row) > 2 else None 
    STATUT = row[33].value if len(row) > 33 else None 
    REP_RDV = row[16].value if len(row) > 16 else None 
    DATE_PRISE_RDV = row[15].value if len(row) > 15 else None 
    CMT_RDV = row[17].value if len(row) > 17 else None 
    DATE_PRISE_RDV, DATE_FIN_TRV = verif_dates(OPENING_DATE_SUR_TIMOS, DATE_PRISE_RDV, DATE_FIN_TRV)
    DATE_FIN_TRV = verif_date(DATE_FIN_TRV)
    DATE_PRISE_RDV= verif_date(DATE_PRISE_RDV)

    if len(row) > 2:
        latitude = clean_coordinate(row[5].value)
        longitude = clean_coordinate(row[6].value)
        if(latitude == '0' and longitude != '0'):
            part1, part2 = separate_numbers(longitude)

        else :
            part1, part2 = separate_numbers(latitude)


        if(part2 is not None):
             if (7 <= float(part1) <= 11):
                longitude = part1.split('_')
                latitude = part2
                if isinstance(longitude, list) and len(longitude) == 1 :
                    longitude=longitude[0]
             else:
                longitude = part2
                latitude = part1

        if((longitude == '_one' ) and latitude != '_one') :
            cleaned_latitude, latitude_result = check_underscores(latitude)
            if (7 <= float(cleaned_latitude) <= 11):
                longitude = cleaned_latitude
                latitude = latitude_result
            else:
                longitude = latitude_result
                latitude = cleaned_latitude
        cleaned_latitude, latitude_result = check_underscores(latitude)
        cleaned_longitude, longitude_result = check_underscores(longitude)
    else:  
        coordinate = clean_coordinate(row[5].value)
        cleaned_latitude, cleaned_longitude = check_underscores(coordinate)

    try:

        cleaned_latitude = safe_float(cleaned_latitude)
        cleaned_longitude = safe_float(cleaned_longitude)
    except (ValueError, TypeError):
        print(f"Erreur de conversion pour le cas CRM {crm_case}: latitude ou longitude invalide")
        continue

    if(crm_case_cleaned != None or verifSTT != None or delegation_clean != None or gouvernorat_clean != None ):
            try:
                  validated_latitude, validated_longitude = validate_coordinates(cleaned_latitude, cleaned_longitude)
                  cleaned_data.append({
                  'CRM_CASE': crm_case_cleaned,
                  'LATITUDE_SITE': validated_latitude,
                  'LONGITUDE_SITE': validated_longitude ,
                  'DATE_CREATION_CRM': DATE_CREATION_CRM,
                  'MSISDN': MSISDN,
                  'CONTACT1_CLIENT': CONTACT1_CLIENT,
                  'CONTACT2_CLIENT': CONTACT2_CLIENT,
                  'CLIENT': client_nettoyé,
                  'REP_TRAVAUX_STT': REP_TRAVAUX_STT,
                  'STT': STT,
                  'Delegation': delegation,
                  'Gouvernorat': gouvernorat,
                  'DATE_AFFECTATION_STT': dateAff_STT,
                  'DES_PACK': pack,
                  'OPENING_DATE_SUR_TIMOS': OPENING_DATE_SUR_TIMOS,
                  'STATUT': STATUT,
                  'REP_RDV': REP_RDV,
                  'DATE_PRISE_RDV': DATE_PRISE_RDV,
                  'CMT_RDV':CMT_RDV,
                  'DATE_FIN_TRV':DATE_FIN_TRV,
                  'Description': description,
                  'Detail':Detail,
                  'entite':entite
                   })
            except Exception as e:
                   print(f"Erreur de validation pour le cas CRM {crm_case}: {e}")
    else:
        Erroned_data.append({
                  'CRM_CASE': crm_case,
                  'LATITUDE_SITE': LATITUDE_SITE_init,
                  'LONGITUDE_SITE': LONGITUDE_SITE_init ,
                  'DATE_CREATION_CRM': DATE_CREATION_CRM,
                  'MSISDN': MSISDN,
                  'CONTACT1_CLIENT': CONTACT1_CLIENT,
                  'CONTACT2_CLIENT': CONTACT2_CLIENT,
                  'CLIENT': client_nettoyé,
                  'REP_TRAVAUX_STT': REP_TRAVAUX_STT,
                  'STT': STT,
                  'Delegation': delegation,
                  'Gouvernorat': gouvernorat,
                  'DATE_AFFECTATION_STT': dateAff_STT,
                  'DES_PACK': pack,
                  'OPENING_DATE_SUR_TIMOS': OPENING_DATE_SUR_TIMOS,
                  'STATUT': STATUT,
                  'REP_RDV':REP_RDV,
                  'DATE_PRISE_RDV': DATE_PRISE_RDV,
                  'CMT_RDV': CMT_RDV,
                  'DATE_FIN_TRV':DATE_FIN_TRV,
                  'Description': description,
                  'Detail':Detail,
                  'entite':entite

                   })
  

new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active
new_sheet.title = "Cleaned Data"

new_workbookError = openpyxl.Workbook()
new_sheetError = new_workbookError.active
new_sheetError.title = "Erroned Data"

headers = ['CRM_CASE', 'DATE_CREATION_CRM', 'LATITUDE_SITE', 'LONGITUDE_SITE', 'MSISDN', 'CONTACT1_CLIENT','CONTACT2_CLIENT', 'CLIENT','REP_TRAVAUX_STT','STT', 'Delegation', ' Gouvernorat' , 'DATE_AFFECTATION_STT ','DES_PACK',
           'OPENING_DATE_SUR_TIMOS','REP_RDV','DATE_PRISE_RDV','CMT_RDV','STATUT','DATE_FIN_TRV','Description','Detail','entite']
new_sheet.append(headers)
new_sheetError.append(headers)

for data in Erroned_data:
    new_sheetError.append([data['CRM_CASE'], data['DATE_CREATION_CRM'], data['LATITUDE_SITE'], 
                           data['LONGITUDE_SITE'], data['MSISDN'], data['CONTACT1_CLIENT'],data['CONTACT2_CLIENT'], data['CLIENT'],data['REP_TRAVAUX_STT'], data['STT'] , data['Delegation']
                           ,data['Gouvernorat'],data['DATE_AFFECTATION_STT'],data['DES_PACK'],data['OPENING_DATE_SUR_TIMOS'],data['REP_RDV'],data['DATE_PRISE_RDV'],data['CMT_RDV']
                          ,data['STATUT'],data['DATE_FIN_TRV'],data['Description'],data['Detail'],data['entite']])

try:
    new_file_pathError = './Erroned data/Erroned_data_resiliation.xlsx'
    new_workbookError.save(new_file_pathError)
    print(f"Les données erronées ont été enregistrées dans : {new_file_pathError}")
except Exception as e:
    print(f"Erreur lors de l'enregistrement des données erronées : {e}")

try:
    new_file_path = './cleaned data/cleaned_data_resiliation.xlsx'
    new_workbook.save(new_file_path)
    print(f"Les données nettoyées ont été enregistrées dans : {new_file_path}")
except Exception as e:
    print(f"Erreur lors de l'enregistrement des données nettoyées : {e}")


connection, cursor = connect_to_mysql()
if not connection or not cursor:
    print("❌ Impossible de se connecter à la base de données")
else:
    try:
        new_file_pathError = './Erroned data/Erroned_data_resiliation.xlsx'
        new_workbookError.save(new_file_pathError)
        print(f"Les données erronées ont été enregistrées dans : {new_file_pathError}")

        new_file_path = './cleaned data/cleaned_data_resiliation.xlsx'
        new_workbook.save(new_file_path)
        print(f"Les données nettoyées ont été enregistrées dans : {new_file_path}")

        insert_query = """
        INSERT INTO resiliation (
            CRM_CASE, DATE_CREATION_CRM, LATITUDE_SITE, LONGITUDE_SITE,
            MSISDN, CONTACT1_CLIENT, CONTACT2_CLIENT, CLIENT,
            REP_TRAVAUX_STT, STT, Delegation, Gouvernorat,
            DATE_AFFECTATION_STT, DES_PACK, OPENING_DATE_SUR_TIMOS,
            STATUT, REP_RDV, DATE_PRISE_RDV, CMT_RDV, DATE_FIN_TRV,
            Description, Detail, entite
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        cleaned_data_unique = remove_duplicates(cleaned_data)

        values_list = []
        for data in cleaned_data_unique:
            values = (
                data.get('CRM_CASE'),
                parse_date(data.get('DATE_CREATION_CRM')),
                safe_float(data.get('LATITUDE_SITE')),
                safe_float(data.get('LONGITUDE_SITE')),
                data.get('MSISDN'),
                data.get('CONTACT1_CLIENT'),
                data.get('CONTACT2_CLIENT'),
                data.get('CLIENT'),
                data.get('REP_TRAVAUX_STT'),
                data.get('STT'),
                data.get('Delegation'),
                data.get('Gouvernorat'),
                parse_date(data.get('DATE_AFFECTATION_STT')),
                data.get('DES_PACK'),
                parse_date(data.get('OPENING_DATE_SUR_TIMOS')),
                data.get('STATUT'),
                data.get('REP_RDV'),
                parse_date(data.get('DATE_PRISE_RDV')),
                data.get('CMT_RDV'),
                parse_date(data.get('DATE_FIN_TRV')),
                data.get('Description'),
                data.get('Detail'),
                data.get('entite')
            )
            values_list.append(values)

        try:
            cursor.executemany(insert_query, values_list)
            connection.commit()
            print(f"✅ {len(values_list)} lignes insérées avec succès !")
        except Error as e:
            print(f"❌ Erreur lors de l'insertion : {e}")
            connection.rollback()

    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()
            print("🔌 Connexion à MySQL fermée.")